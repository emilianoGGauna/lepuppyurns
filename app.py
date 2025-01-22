from flask import Flask, request, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask import Flask, render_template, request, send_file, flash, redirect, url_for
from tempfile import NamedTemporaryFile
from functools import wraps
from flask import session, redirect, url_for, flash
import os
import logging
from pymongo import MongoClient
from dotenv import load_dotenv
from PasswordManager import PasswordManager
import hashlib
from datetime import datetime
import urllib.parse
import pandas as pd
import random
import string
import base64
from jinja2 import TemplateNotFound
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from bson import ObjectId  # Importar para manejar ObjectId en MongoDB
from flask import Flask, render_template, redirect, url_for, flash
from datetime import datetime, timedelta
import plotly.graph_objs as go
from plotly.utils import PlotlyJSONEncoder
import json
from pymongo import MongoClient
import logging
import re

# Configurar logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger()

# Cargar variables de entorno desde .env
load_dotenv()
MONGO_URI = os.getenv("MONGO_URI")
APP_KEY = os.getenv("APP_KEY")

if not MONGO_URI:
    logger.error("La variable MONGO_URI no está configurada en el archivo .env")
    raise ValueError("La variable MONGO_URI no está configurada en el archivo .env")

# Configuración de Flask
app = Flask(__name__, static_folder="static", template_folder="templates")
app.secret_key = APP_KEY

# Conexión a la base de datos MongoDB
try:
    client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=30000)

    db = client["lepuppy"]
    usuarios_collection = db["usuarios"]
    catalogo_collection = db["catalogo"]
    logger.info("Conexión a la base de datos establecida correctamente.")
except Exception as e:
    logger.error(f"Error al conectar con la base de datos MongoDB: {e}")
    raise

# Instanciar PasswordManager
password_manager = PasswordManager()

# Middleware para verificar roles
def role_required(role):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            if "user_role" not in session or session["user_role"] != role:
                flash("Acceso denegado.")
                return redirect(url_for("login"))
            return func(*args, **kwargs)
        return wrapper
    return decorator

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        identifier = request.form.get("email")  # Puede ser correo o teléfono
        password = request.form.get("password")

        if not identifier or not password:
            flash("Por favor, complete todos los campos.")
            return redirect(url_for("login"))

        try:
            contraseña_encriptada_ingresada = password_manager.encrypt_password(password)

            # Verificar si el identificador es un correo o un teléfono
            if "@" in identifier:
                # Es un correo
                usuario = usuarios_collection.find_one({
                    "email": identifier,
                    "contraseña": contraseña_encriptada_ingresada
                })
            else:
                # Es un número de teléfono
                usuario = usuarios_collection.find_one({
                    "telefono": identifier,
                    "contraseña": contraseña_encriptada_ingresada
                })
        except Exception as e:
            flash("Error al conectar con la base de datos.")
            logger.error(f"Error: {e}")
            return redirect(url_for("login"))

        if usuario:
            session["user_role"] = usuario.get("access")
            session["user_name"] = usuario.get("client_name", "Usuario")
            session["user_id"] = str(usuario["_id"])  # Asegúrate de convertir el ID a string
            if usuario["access"] == "admin":
                return redirect(url_for("adminplatform"))
            elif usuario["access"] == "cliente":
                return redirect(url_for("clientecatalogo"))
        else:
            flash("Usuario o contraseña incorrectos.")
            return redirect(url_for("login"))
    return render_template("login.html")
    
@app.route("/olvide-contraseña", methods=["GET", "POST"])
def olvide_contraseña():
    if request.method == "POST":
        email = request.form.get("email")
        phone = request.form.get("phone")
        new_password = request.form.get("new_password")
        confirm_password = request.form.get("confirm_password")

        # Validar que las contraseñas coincidan
        if new_password != confirm_password:
            flash("Las contraseñas no coinciden.")
            return redirect(url_for("olvide_contraseña"))

        # Verificar si el usuario con el correo y teléfono existe
        usuario = usuarios_collection.find_one({"email": email, "phone": phone})
        if not usuario:
            flash("No se encontró una cuenta con este correo electrónico y número de teléfono.")
            return redirect(url_for("olvide_contraseña"))

        # Encriptar la nueva contraseña usando PasswordManager
        try:
            hashed_password = password_manager.encrypt_password(new_password)
        except Exception as e:
            flash("Error al encriptar la nueva contraseña.")
            logger.error(f"Error al encriptar la contraseña: {e}")
            return redirect(url_for("olvide_contraseña"))

        # Actualizar la contraseña en la base de datos
        try:
            usuarios_collection.update_one(
                {"email": email, "phone": phone},
                {"$set": {"contraseña": hashed_password}}
            )
            flash("Tu contraseña ha sido actualizada exitosamente.")
            return redirect(url_for("login"))
        except Exception as e:
            flash("Error al actualizar la contraseña.")
            logger.error(f"Error al actualizar la contraseña: {e}")
            return redirect(url_for("olvide_contraseña"))

    return render_template("olvide_contraseña.html")


@app.route("/registro", methods=["GET", "POST"])
def registro():
    if request.method == "POST":
        try:
            # Extract form data
            client_name = request.form.get("client_name")
            phone = request.form.get("phone")
            email = request.form.get("email")
            confirm_email = request.form.get("confirm_email")  # Nuevo campo para confirmar correo
            password = request.form.get("password")
            confirm_password = request.form.get("confirm_password")

            # Validate email and confirm email
            if email != confirm_email:
                flash("Los correos electrónicos no coinciden.")
                return redirect(url_for("registro"))

            # Validate passwords
            if password != confirm_password:
                flash("Las contraseñas no coinciden.")
                return redirect(url_for("registro"))

            # Encrypt password
            encrypted_password = password_manager.encrypt_password(password)

            # Create user data with fixed access type "cliente"
            user_data = {
                "client_name": client_name,
                "phone": phone,
                "email": email,
                "access": "cliente",
                "contraseña": encrypted_password,
            }

            # Generate client_id as MD5 hash of user data
            user_json_string = f"{client_name}{phone}{email}cliente".encode("utf-8")
            client_id = hashlib.md5(user_json_string).hexdigest()
            user_data["client_id"] = client_id

            # Insert into MongoDB
            usuarios_collection.insert_one(user_data)
            flash("Registro exitoso. Ahora puedes iniciar sesión.")
            return redirect(url_for("login"))
        except Exception as e:
            flash("Error al registrar usuario.")
            logger.error(f"Error al registrar usuario: {e}")
            return redirect(url_for("registro"))

    return render_template("registro.html")


##------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
##------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
## CLient PLATFORM
##------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
##------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
@app.route("/cart_count", methods=["GET"])
def cart_count():
    if "user_id" not in session:
        return jsonify({"count": 0})  # Si no hay sesión, el carrito está vacío

    try:
        client_id = session.get("user_id")
        carrito_collection = db["carrito"]
        count = carrito_collection.count_documents({"client_id": client_id})  # Conteo de productos
        return jsonify({"count": count})
    except Exception as e:
        logger.error(f"Error al obtener la cantidad de artículos en el carrito: {e}")
        return jsonify({"count": 0})

@app.route("/clientecatalogo")
def clientecatalogo():
    try:
        # Consultar catálogo ordenado por sort_order
        catalogos = list(
            catalogo_collection.find(
                {}, {"_id": 0, "Tipo de Modelo": 1, "img": 1, "model-uuid": 1, "sort_order": 1}
            ).sort("sort_order", 1)
        )
        logger.info(f"Catálogo cargado correctamente")
        return render_template("cliente/clientecatalogo.html", catalogos=catalogos)
    except Exception as e:
        flash(f"Error al cargar el catálogo.")
        logger.error(f"Error al cargar el catálogo: {e}")
        return redirect(url_for("login"))


@app.route("/clienteforms/<model_uuid>", methods=["GET", "POST"])
def clienteforms(model_uuid):
    if request.method == "GET":
        try:
            logger.info(f"Buscando modelo con UUID: {model_uuid}")
            
            # Buscar modelo en la base de datos
            model = catalogo_collection.find_one({"model-uuid": model_uuid}, {"_id": 0})
            
            if model:
                logger.info(f"Modelo encontrado: {model.get('Tipo de Modelo', 'Desconocido')}")

                # Validar si el modelo tiene 'Forms' y manejar el caso de que falte 'Tipo Urna'
                forms = model.get("Forms", {})
                if not isinstance(forms, dict):
                    logger.warning(f"'Forms' no es válido para el modelo con UUID: {model_uuid}")
                    model["Forms"] = {}

                if "Tipo Urna" not in model["Forms"]:
                    logger.warning(f"'Tipo Urna' no encontrado en el modelo con UUID: {model_uuid}")
                    model["Forms"]["Tipo Urna"] = {}  # Valor por defecto vacío

                return render_template("cliente/clienteforms.html", model=model)

            else:
                flash("Modelo no encontrado.")
                logger.warning(f"Modelo no encontrado para UUID: {model_uuid}")
                return redirect(url_for("clientecatalogo"))

        except Exception as e:
            flash("Error al cargar el formulario. Intente nuevamente.")
            logger.error(f"Error al cargar el formulario para UUID {model_uuid}: {e}")
            return redirect(url_for("clientecatalogo"))

@app.route("/add_to_cart/<model_uuid>", methods=["POST"])
def add_to_cart(model_uuid):
    if "user_id" not in session:  # Verifica si el usuario está autenticado
        return jsonify({"success": False, "error": "Por favor, inicie sesión para añadir al carrito."})

    try:
        # Obtener datos del formulario
        cantidad = request.form.get("cantidad", type=int)
        client_id = session.get("user_id")  # Recuperar el ID del cliente desde la sesión
        model = catalogo_collection.find_one({"model-uuid": model_uuid}, {"_id": 0})
        forms_data = {key: value for key, value in request.form.items() if key != "cantidad"}

        if not model:
            return jsonify({"success": False, "error": "Modelo no encontrado."})

        # Crear un identificador único para la combinación de opciones
        forms_hash = hashlib.sha256(str(forms_data).encode()).hexdigest()

        # Crear/Actualizar el carrito en MongoDB
        carrito_collection = db["carrito"]
        carrito_collection.update_one(
            {"client_id": client_id, "model_uuid": model_uuid, "forms_hash": forms_hash},
            {"$set": {
                "model": model,
                "client_id": client_id,
                "model_uuid": model_uuid,
                "cantidad": cantidad,
                "forms_data": forms_data,
                "forms_hash": forms_hash  # Almacenar el hash en la BD
            }},
            upsert=True
        )
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"Error al añadir al carrito: {e}")
        return jsonify({"success": False, "error": "Ocurrió un error al añadir al carrito."})



@app.route("/clientecarrito")
def clientecarrito():
    if "user_id" not in session:  # Asegurarse de que el usuario esté autenticado
        flash("Por favor, inicie sesión para ver su carrito.")
        return redirect(url_for("login"))

    try:
        # Recuperar el ID del cliente desde la sesión
        client_id = session.get("user_id")
        
        # Consultar los objetos del carrito de este cliente
        carrito_collection = db["carrito"]
        carrito_items = list(carrito_collection.find({"client_id": client_id}, {"_id": 0}))

        if not carrito_items:
            flash("Su carrito está vacío.")
            return render_template("cliente/clientecarrito.html", carrito=[])

        # Renderizar la plantilla del carrito
        return render_template("cliente/clientecarrito.html", carrito=carrito_items)
    except Exception as e:
        logger.error(f"Error al cargar el carrito: {e}")
        flash("Ocurrió un error al cargar su carrito.")
        return redirect(url_for("clientecatalogo"))
    
@app.route("/update_cart", methods=["POST"])
def update_cart():
    if "user_id" not in session:
        flash("Por favor, inicie sesión para editar su carrito.")
        return redirect(url_for("login"))

    try:
        # Recuperar datos del formulario
        client_id = session.get("user_id")
        model_uuid = request.form.get("model_uuid")
        forms_hash = request.form.get("forms_hash")  # Identificador único del carrito
        new_quantity = int(request.form.get("cantidad"))

        # Validar que la cantidad sea positiva
        if new_quantity < 1:
            flash("La cantidad debe ser mayor a 0.")
            return redirect(url_for("clientecarrito"))

        carrito_collection = db["carrito"]
        result = carrito_collection.update_one(
            {"client_id": client_id, "model_uuid": model_uuid, "forms_hash": forms_hash},
            {"$set": {"cantidad": new_quantity}}
        )

        if result.matched_count > 0:
            flash("Cantidad actualizada exitosamente.")
        else:
            flash("No se encontró el producto en el carrito.")

    except Exception as e:
        logger.error(f"Error al actualizar el carrito: {e}")
        flash("Ocurrió un error al actualizar su carrito.")

    return redirect(url_for("clientecarrito"))


@app.route("/delete_from_cart", methods=["POST"])
def delete_from_cart():
    if "user_id" not in session:
        flash("Por favor, inicie sesión para editar su carrito.")
        return redirect(url_for("login"))

    try:
        client_id = session.get("user_id")
        model_uuid = request.form.get("model_uuid")

        # Debug log to ensure correct model_uuid is received
        logger.info(f"Attempting to delete item with model_uuid: {model_uuid} for client_id: {client_id}")

        # Ensure model_uuid exists in the form
        if not model_uuid:
            flash("Error: Producto no encontrado.")
            return redirect(url_for("clientecarrito"))

        carrito_collection = db["carrito"]
        result = carrito_collection.delete_one({"client_id": client_id, "model_uuid": model_uuid})

        if result.deleted_count > 0:
            flash("Producto eliminado del carrito.")
        else:
            flash("El producto no pudo ser eliminado. Verifique nuevamente.")
    except Exception as e:
        logger.error(f"Error al eliminar producto del carrito: {e}")
        flash("Ocurrió un error al eliminar el producto.")

    return redirect(url_for("clientecarrito"))

@app.route("/finalizar-compra", methods=["POST"])
def finalizar_compra():
    if "user_id" not in session:
        flash("Por favor, inicie sesión para completar su compra.")
        return redirect(url_for("login"))

    try:
        client_id = session.get("user_id")
        carrito_collection = db["carrito"]
        pedidos_collection = db["pedidos"]
        usuarios_collection = db["usuarios"]
        catalogo_collection = db["catalogo"]

        # Verificar si la colección 'pedidos' existe, si no, crearla
        if "pedidos" not in db.list_collection_names():
            pedidos_collection = db.create_collection("pedidos")
            logger.info("La colección 'pedidos' no existía y ha sido creada.")

        # Fetch the cart for the user
        orden = list(carrito_collection.find({"client_id": client_id}))

        if not orden:
            flash("No hay productos en tu carrito.")
            return redirect(url_for("clientecarrito"))

        # Convertir client_id a ObjectId
        try:
            client_object_id = ObjectId(client_id)
        except Exception as e:
            logger.error(f"Error al convertir client_id a ObjectId: {e}")
            flash("Error interno. Contacta al soporte.")
            return redirect(url_for("clientecarrito"))

        # Fetch user details
        cliente_info = usuarios_collection.find_one({"_id": client_object_id})
        if not cliente_info:
            logger.error(f"No se encontró un usuario con ID: {client_id}")
            flash("Error al procesar tu información de cliente.")
            return redirect(url_for("clientecarrito"))

        client_name = cliente_info.get("client_name", "Cliente Desconocido")

        # Prepare the order
        time_stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        orden_id = f"ORD-{datetime.now().strftime('%Y%m%d%H%M%S')}"

        pedidos = []
        total_urnas = 0
        for item in orden:
            # Get the model name from the catalog
            model_uuid = item["model_uuid"]
            modelo = catalogo_collection.find_one({"model-uuid": model_uuid})
            nombre_modelo = modelo.get("Tipo de Modelo", "Modelo Desconocido") if modelo else "Modelo Desconocido"
            
            cantidad = item["cantidad"]
            
            total_urnas += cantidad
            pedidos.append({
                "model_uuid": model_uuid,
                "modelo": nombre_modelo,
                "cantidad": item["cantidad"],
                "atributos": item["forms_data"]
            })

        # Create the order
        order_data = {
            "orden-id": orden_id,
            "cliente-id": client_id,
            "cliente-nombre": client_name,
            "time-stamp": time_stamp,
            "Estado": "En proceso",
            "pedidos": pedidos
        }

        # Insert order into pedidos collection
        pedidos_collection.insert_one(order_data)

        # Clear the cart
        carrito_collection.delete_many({"client_id": client_id})

        # Prepare WhatsApp message
        mensaje = (
            f"Hola, soy {client_name}.\n"
            f"Orden ID: {orden_id}\n"
            f"Fecha de compra: {time_stamp}\n"
            f"Cantidad de pedidos realizados: {len(pedidos)}\n"
            f"Total de urnas: {total_urnas}\n"
        )

        mensaje_codificado = urllib.parse.quote(mensaje)
        whatsapp_link = f"https://api.whatsapp.com/send?phone=3325648862&text={mensaje_codificado}"

        return redirect(whatsapp_link)

    except Exception as e:
        logger.error(f"Error al finalizar la compra: {e}")
        flash("Ocurrió un error al finalizar tu compra.")
        return redirect(url_for("clientecarrito"))

@app.route("/clientepedidos", methods=["GET"])
def clientepedidos():
    if "user_id" not in session:
        flash("Por favor, inicie sesión para ver sus pedidos.")
        return redirect(url_for("login"))

    try:
        # Recuperar el ID del cliente desde la sesión
        client_id = session.get("user_id")
        
        # Consultar los pedidos del cliente
        pedidos_collection = db["pedidos"]
        pedidos = list(pedidos_collection.find({"cliente-id": client_id}))

        # Dividir los pedidos por estado
        pedidos_en_proceso = [pedido for pedido in pedidos if pedido.get("Estado") == "En proceso"]
        pedidos_terminados = [pedido for pedido in pedidos if pedido.get("Estado") == "Terminado"]

        # Calcular totales para cada orden
        def calcular_totales(orden, color):
            total_urnas = sum(item.get("cantidad", 0) for item in orden.get("pedidos", []))
            pedidos_por_orden = len(orden.get("pedidos", []))
            return {
                "orden_id": orden.get("orden-id"),
                "time_stamp": orden.get("time-stamp"),
                "estado": orden.get("Estado"),
                "cliente_nombre": orden.get("cliente-nombre"),
                "total_urnas": total_urnas,
                "pedidos_por_orden": pedidos_por_orden,
                "detalles": orden.get("pedidos", []),
                "card_color": color
            }

        # Generar detalles de pedidos en proceso y terminados
        detalles_en_proceso = [calcular_totales(orden, "green") for orden in pedidos_en_proceso]
        detalles_terminados = [calcular_totales(orden, "red") for orden in pedidos_terminados]

        return render_template(
            "cliente/clientepedidos.html",
            pedidos_en_proceso=detalles_en_proceso,
            pedidos_terminados=detalles_terminados
        )
    except Exception as e:
        logger.error(f"Error al cargar los pedidos del cliente: {e}")
        flash("Ocurrió un error al cargar sus pedidos.")
        return redirect(url_for("clientecatalogo"))



##------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
##------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
## ADMIN PLATFORM
##------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
##------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def generate_page_one(ws, pedido):
    """Genera la primera página del pedido con estilo mejorado."""
    # Estilo para el encabezado
    bold_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Agregar encabezado general
    ws.merge_cells("A1:F1")
    ws["A1"] = "Resumen del Pedido"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].alignment = center_alignment
    ws["A1"].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    # Calcular la cantidad total y el número de modelos
    total_cantidad = sum(item["cantidad"] for item in pedido["pedidos"])
    total_modelos = len(pedido["pedidos"])

    # Información principal del pedido
    info_pedido = [
        ["Orden ID", pedido.get("orden-id", "-")],
        ["Cliente Nombre", pedido.get("cliente-nombre", "-")],
        ["Fecha", pedido.get("time-stamp", "-")],
        ["Cantidad Total", total_cantidad],
        ["Número de Modelos", total_modelos]
    ]

    row = 2  # Comenzar en la fila 2
    for key, value in info_pedido:
        ws[f"A{row}"] = key
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = value
        row += 1

    # Espaciado entre secciones
    row += 1

    # Crear conjunto único de todos los atributos dinámicos en los modelos
    all_attributes = set()
    for item in pedido["pedidos"]:
        all_attributes.update(item.get("atributos", {}).keys())

    # Encabezados para la tabla
    headers = ["Modelo", "Cantidad"] + sorted(all_attributes)  # Ordenar atributos alfabéticamente
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Agregar filas con datos de los modelos en el pedido
    row += 1
    for item in pedido["pedidos"]:
        data_row = [
            item.get("modelo", "-"),
            item.get("cantidad", 0),
        ]
        # Agregar valores de atributos dinámicos en el orden de los encabezados
        for attr in sorted(all_attributes):
            data_row.append(item.get("atributos", {}).get(attr, "-"))
        for col, value in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
        row += 1

    # Ajustar ancho de columnas automáticamente
    for col_num, column_cells in enumerate(ws.iter_cols(min_row=1, max_row=row, min_col=1, max_col=len(headers)), start=1):
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=10)
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

def generate_page_two(wb, pedido, catalogo_collection):
    """Generates the second page with detailed Corte Láser information."""
    bold_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Create 'Corte Láser' sheet
    ws_laser = wb.create_sheet(title="Corte Láser")

    # Normalize keys (standardize similar attribute names)
    def normalize_key(key):
        return key.strip().lower().replace(" ", "_")

    # Collect all possible attributes from Corte Láser and form attributes
    all_attributes = set()
    for item in pedido["pedidos"]:
        modelo = item.get("modelo", "-")
        catalogo_item = catalogo_collection.find_one({"Tipo de Modelo": modelo})
        if catalogo_item and isinstance(catalogo_item.get("corte_lazer"), dict):
            all_attributes.update(normalize_key(k) for k in catalogo_item["corte_lazer"].keys())
        if isinstance(item.get("atributos"), dict):
            all_attributes.update(normalize_key(k) for k, v in item["atributos"].items() if k.lower() != "color")

    # Ensure `modelo` is included and add it first
    all_attributes = ["modelo"] + [attr for attr in sorted(all_attributes) if attr != "modelo"]

    # Excluir columnas que contengan "color" en cualquier parte del nombre
    # Excluir columnas que contengan "color" o "base" en cualquier parte del nombre
    excluded_columns = {attr for attr in all_attributes if re.search(r"(color|Base)", attr, re.IGNORECASE)}

    # Excluir columnas específicas y asegurar que "tipo-urna" siempre esté al final
    all_attributes = [attr for attr in all_attributes if attr not in excluded_columns and attr != "tipo_urna"] + ["tipo-urna"]

    # Create dynamic headers
    headers_laser = ["Cantidad"] + all_attributes
    for col, header in enumerate(headers_laser, start=1):
        cell = ws_laser.cell(row=1, column=col)
        cell.value = header
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Add rows with data
    row_num = 2
    for item in pedido["pedidos"]:
        cantidad = item.get("cantidad", 0)

        # Look up in catalog using `Tipo de Modelo`
        modelo = item.get("modelo", "-")
        catalogo_item = catalogo_collection.find_one({"Tipo de Modelo": modelo})
        corte_lazer = {normalize_key(k): v for k, v in catalogo_item.get("corte_lazer", {}).items()} if catalogo_item else {}
        atributos_form = {normalize_key(k): v for k, v in item.get("atributos", {}).items() if k.lower() != "color"}

        # Build the row data
        data_row = [cantidad]  # Start with quantity
        for attr in all_attributes:
            if attr == "modelo":
                value = modelo  # Ensure `modelo` is included
            else:
                value = atributos_form.get(attr, corte_lazer.get(attr, "-"))
            data_row.append(value)

        # Write the row to the sheet
        for col, value in enumerate(data_row, start=1):
            cell = ws_laser.cell(row=row_num, column=col)
            cell.value = value

        row_num += 1

    # Remove columns with all null values
    for col in range(len(headers_laser), 0, -1):
        if all(
            ws_laser.cell(row=row, column=col).value in [None, "-", ""] for row in range(2, row_num)
        ):
            ws_laser.delete_cols(col)

    # Merge duplicate rows (except the "Cantidad" column)
    row_data = []
    for row in ws_laser.iter_rows(min_row=2, max_row=row_num - 1, min_col=1, max_col=len(headers_laser)):
        row_values = [cell.value for cell in row]
        row_data.append(row_values)

    unique_rows = {}
    for row in row_data:
        key = tuple(row[1:])  # Exclude "Cantidad" to create the key
        cantidad = int(row[0]) if isinstance(row[0], int) else 0
        if key in unique_rows:
            unique_rows[key] += cantidad
        else:
            unique_rows[key] = cantidad

    # Write unique rows back
    ws_laser.delete_rows(2, ws_laser.max_row)
    row_num = 2
    for key, cantidad in unique_rows.items():
        data_row = [cantidad] + list(key)
        for col, value in enumerate(data_row, start=1):
            cell = ws_laser.cell(row=row_num, column=col)
            cell.value = value
        row_num += 1

    # Auto-adjust column widths
    for col_num, column_cells in enumerate(ws_laser.iter_cols(min_row=1, max_row=row_num - 1, min_col=1, max_col=len(headers_laser)), start=1):
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=10)
        ws_laser.column_dimensions[get_column_letter(col_num)].width = max_length + 2

    # Eliminate duplicate headers
    seen_headers = set()
    duplicate_columns = []
    for col in range(1, len(headers_laser) + 1):
        header_value = ws_laser.cell(row=1, column=col).value
        if header_value in seen_headers:
            duplicate_columns.append(col)
        else:
            seen_headers.add(header_value)

    # Delete duplicate columns
    for col in reversed(duplicate_columns):
        ws_laser.delete_cols(col)

    # Map column names to new names if necessary
    column_mapping = {
        "¿quieres_el_logo_de_tu_empresa?": "logo_grabado",
        "tipo-urna": "tipo-madera"
    }

    for col in range(1, ws_laser.max_column + 1):
        cell_value = ws_laser.cell(row=1, column=col).value
        if cell_value in column_mapping:
            ws_laser.cell(row=1, column=col).value = column_mapping[cell_value]

    # Auto-adjust column widths and enable text wrapping
    for col_num, column_cells in enumerate(
        ws_laser.iter_cols(min_row=1, max_row=row_num - 1, min_col=1, max_col=ws_laser.max_column), start=1
    ):
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=10)
        column_letter = get_column_letter(col_num)
        ws_laser.column_dimensions[column_letter].width = max_length + 2

        for cell in column_cells:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


    # Adjust row heights for better readability
    for row in ws_laser.iter_rows(min_row=1, max_row=row_num - 1, min_col=1, max_col=ws_laser.max_column):
        for cell in row:
            if cell.value:
                ws_laser.row_dimensions[cell.row].height = 15

    cantidad_header = ws_laser.cell(row=1, column=1).value
    modelo_header = ws_laser.cell(row=1, column=2).value

    # Asegurarnos de que los nombres se correspondan antes de cambiar
    if cantidad_header.lower() == "cantidad" and modelo_header.lower() == "modelo":
        ws_laser.cell(row=1, column=1).value = "Modelo"  # Cambiar "Cantidad" por "Modelo"
        ws_laser.cell(row=1, column=2).value = "Cantidad"  # Cambiar "Modelo" por "Cantidad"

    # Cambiar los valores de las dos primeras columnas en todas las filas
    for row in ws_laser.iter_rows(min_row=2, max_row=row_num - 1, min_col=1, max_col=ws_laser.max_column):
        first_value = row[0].value
        second_value = row[1].value
        # Intercambiar valores
        row[0].value = second_value
        row[1].value = first_value

                
@app.route("/export_pedido/<orden_id>", methods=["GET"])
def export_pedido(orden_id):
    try:
        pedidos_collection = db["pedidos"]
        catalogo_collection = db["catalogo"]
        pedido = pedidos_collection.find_one({"orden-id": orden_id})

        if not pedido:
            flash("Pedido no encontrado.")
            return redirect(url_for("adminplatform"))

        # Crear un nuevo archivo Excel con openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen del Pedido"

        # Generar página 1
        generate_page_one(ws, pedido)

        # Generar página 2
        generate_page_two(wb, pedido, catalogo_collection)

        # Guardar el archivo temporalmente
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            wb.save(temp_file.name)
            temp_file_path = temp_file.name

        # Descargar el archivo
        response = send_file(
            temp_file_path,
            as_attachment=True,
            download_name=f"pedido_{orden_id}.xlsx"
        )
        response.call_on_close(lambda: os.unlink(temp_file_path))  # Eliminar archivo al terminar
        return response

    except Exception as e:
        logger.error(f"Error al exportar pedido {orden_id}: {e}")
        flash("Ocurrió un error al exportar el pedido.")
        return redirect(url_for("adminplatform"))


@app.route("/update_status/<orden_id>", methods=["POST"])
def update_status(orden_id):
    """Actualizar el estado de un pedido."""
    try:
        new_status = request.form["new_status"]
        pedidos_collection = db["pedidos"]
        pedidos_collection.update_one(
            {"orden-id": orden_id},
            {"$set": {"Estado": new_status}}
        )
        flash(f"Estado del pedido {orden_id} actualizado a {new_status}.")
    except Exception as e:
        logger.error(f"Error al actualizar el estado del pedido {orden_id}: {e}")
        flash("Ocurrió un error al actualizar el estado del pedido.")
    return redirect(url_for("adminplatform"))

@app.route("/test_pedidos", methods=["GET"])
def test_pedidos():
    """Ruta para probar los datos de pedidos en JSON."""
    pedidos_collection = db["pedidos"]
    pedidos = list(pedidos_collection.find({}))
    return jsonify(pedidos)


@app.route("/delete_pedido/<orden_id>", methods=["POST"])
def delete_pedido(orden_id):
    """
    Elimina un pedido basado en su Orden ID.
    """
    try:
        pedidos_collection = db["pedidos"]
        result = pedidos_collection.delete_one({"orden-id": orden_id})

        if result.deleted_count > 0:
            flash(f"Pedido {orden_id} eliminado exitosamente.")
        else:
            flash(f"No se encontró el pedido con Orden ID {orden_id}.")
    except Exception as e:
        logger.error(f"Error al eliminar el pedido {orden_id}: {e}")
        flash("Ocurrió un error al intentar eliminar el pedido.")
    
    return redirect(url_for("adminplatform"))

@app.route("/adminplatform", methods=["GET"])
def adminplatform():
    try:
        pedidos_collection = db["pedidos"]
        pedidos = list(pedidos_collection.find({}))

        # Dividir pedidos por estado
        pedidos_en_proceso = []
        pedidos_terminados = []

        for pedido in pedidos:
            total_cantidad = sum(item["cantidad"] for item in pedido["pedidos"])
            total_pedidos = len(pedido["pedidos"])
            pedido_data = {
                "Orden ID": pedido.get("orden-id", "-"),
                "Cliente Nombre": pedido.get("cliente-nombre", "-"),
                "Fecha": pedido.get("time-stamp", "-"),
                "Estado": pedido.get("Estado", "-")
            }
            if pedido_data["Estado"] == "En proceso":
                pedidos_en_proceso.append(pedido_data)
            elif pedido_data["Estado"] == "Terminado":
                pedidos_terminados.append(pedido_data)

        # Ordenar por fecha de mayor a menor
        pedidos_en_proceso.sort(key=lambda x: datetime.strptime(x["Fecha"], "%Y-%m-%d %H:%M:%S"), reverse=True)
        pedidos_terminados.sort(key=lambda x: datetime.strptime(x["Fecha"], "%Y-%m-%d %H:%M:%S"), reverse=True)

        # Limitar a los últimos 5 pedidos terminados
        pedidos_terminados_mostrados = pedidos_terminados[:5]

        # Totales
        total_en_proceso = len(pedidos_en_proceso)
        total_terminados = len(pedidos_terminados)

        return render_template(
            "admin/adminplatform.html",
            pedidos_en_proceso=pedidos_en_proceso,
            pedidos_terminados=pedidos_terminados_mostrados,
            total_en_proceso=total_en_proceso,
            total_terminados=total_terminados
        )

    except Exception as e:
        logger.error(f"Error en adminplatform: {e}")
        flash("Ocurrió un error al cargar los pedidos.")
        return redirect(url_for("login"))

    
@app.route("/adminusuarios")
def adminusuarios():
    try:
        # Fetch all user data, including client_id
        usuarios = list(
            usuarios_collection.find({}, {"_id": 0, "email": 1, "client_name": 1, "access": 1, "client_id": 1})
        )
        logger.info("Usuarios cargados correctamente.")
        return render_template("admin/adminusuarios.html", usuarios=usuarios)
    except Exception as e:
        flash("Error al cargar los usuarios.")
        logger.error(f"Error al cargar los usuarios: {e}")
        return redirect(url_for("login"))


@app.route("/adduser", methods=["POST"])
def adduser():
    try:
        # Extract form data
        client_name = request.form.get("client_name")
        phone = request.form.get("phone")
        email = request.form.get("email")
        access = request.form.get("access")
        password = request.form.get("password")
        confirm_password = request.form.get("confirm_password")

        # Validate passwords
        if password != confirm_password:
            flash("Las contraseñas no coinciden.")
            logger.warning("Intento fallido de creación de usuario: las contraseñas no coinciden.")
            return redirect(url_for("adminusuarios"))

        # Encrypt password
        encrypted_password = password_manager.encrypt_password(password)

        # Create user data
        user_data = {
            "client_name": client_name,
            "phone": phone,
            "email": email,
            "access": access,
            "contraseña": encrypted_password,
        }

        # Generate client_id as MD5 hash of user data
        user_json_string = f"{client_name}{phone}{email}{access}".encode("utf-8")
        client_id = hashlib.md5(user_json_string).hexdigest()
        user_data["client_id"] = client_id

        # Insert into MongoDB
        usuarios_collection.insert_one(user_data)
        flash("Usuario añadido exitosamente.")
        logger.info(f"Usuario añadido: {email}, client_id: {client_id}")

        return redirect(url_for("adminusuarios"))
    except Exception as e:
        flash("Error al añadir usuario.")
        logger.error(f"Error al añadir usuario: {e}")
        return redirect(url_for("adminusuarios"))


@app.route("/deleteuser", methods=["POST"])
def deleteuser():
    try:
        # Get client_id from form data
        client_id = request.form.get("client_id")
        logger.info(f"Received client_id for deletion: {client_id}")
        if not client_id:
            flash("ID del cliente no proporcionado.")
            logger.warning("Intento de eliminación fallido: client_id no proporcionado.")
            return redirect(url_for("adminusuarios"))

        # Attempt to delete the user
        result = usuarios_collection.delete_one({"client_id": client_id})
        if result.deleted_count > 0:
            flash("Usuario eliminado exitosamente.")
            logger.info(f"Usuario eliminado con client_id: {client_id}")
        else:
            flash("Usuario no encontrado.")
            logger.warning(f"No se encontró el usuario con client_id: {client_id}")

        return redirect(url_for("adminusuarios"))
    except Exception as e:
        flash("Error al eliminar el usuario.")
        logger.error(f"Error al eliminar usuario: {e}")
        return redirect(url_for("adminusuarios"))

@app.route("/update_sort_order/<model_uuid>", methods=["POST"])
def update_sort_order(model_uuid):
    try:
        data = request.json
        new_sort_order = data.get("sort_order")

        if new_sort_order is None:
            return jsonify({"error": "El número de orden es requerido."}), 400

        # Obtener el modelo actual
        current_item = catalogo_collection.find_one({"model-uuid": model_uuid})
        if not current_item:
            return jsonify({"error": "Modelo no encontrado."}), 404

        # Ajustar el sort_order de los demás modelos
        catalogo_collection.update_many(
            {"sort_order": {"$gte": new_sort_order}}, 
            {"$inc": {"sort_order": 1}}
        )

        # Actualizar el sort_order del modelo seleccionado
        catalogo_collection.update_one(
            {"model-uuid": model_uuid},
            {"$set": {"sort_order": new_sort_order}}
        )

        # Reorganizar `sort_order` para eliminar huecos
        items = list(
            catalogo_collection.find({}, {"model-uuid": 1}).sort("sort_order", 1)
        )
        for idx, item in enumerate(items):
            catalogo_collection.update_one(
                {"model-uuid": item["model-uuid"]},
                {"$set": {"sort_order": idx + 1}}
            )

        return jsonify({"message": "Orden actualizado correctamente."}), 200
    except Exception as e:
        logger.error(f"Error al actualizar el sort_order: {e}")
        return jsonify({"error": "Error interno del servidor."}), 500

@app.route("/change_image/<model_uuid>", methods=["POST"])
def change_image(model_uuid):
    try:
        # Validar si se envió un archivo
        if 'new_image' not in request.files:
            return jsonify({"error": "No se envió ninguna imagen."}), 400

        new_image = request.files['new_image']

        # Validar que el archivo sea una imagen
        if not new_image or new_image.filename == '':
            return jsonify({"error": "Archivo inválido."}), 400

        # Convertir la imagen a Base64
        new_image_base64 = base64.b64encode(new_image.read()).decode("utf-8")

        # Actualizar la base de datos con la nueva imagen
        result = catalogo_collection.update_one(
            {"model-uuid": model_uuid},
            {"$set": {"img.modelos": new_image_base64}}
        )

        if result.matched_count == 0:
            return jsonify({"error": "Modelo no encontrado."}), 404

        return jsonify({"message": "Imagen cambiada con éxito."}), 200

    except Exception as e:
        logger.error(f"Error al cambiar la imagen para {model_uuid}: {e}")
        return jsonify({"error": "Error interno del servidor."}), 500

@app.route("/change_description_image/<model_uuid>", methods=["POST"])
def change_description_image(model_uuid):
    try:
        # Validar si se envió un archivo
        if 'new_image' not in request.files:
            return jsonify({"error": "No se envió ninguna imagen."}), 400

        new_image = request.files['new_image']

        # Validar que el archivo sea una imagen
        if not new_image or new_image.filename == '':
            return jsonify({"error": "Archivo inválido."}), 400

        # Convertir la imagen a Base64
        new_image_base64 = base64.b64encode(new_image.read()).decode("utf-8")

        # Actualizar la base de datos con la nueva imagen descriptiva
        result = catalogo_collection.update_one(
            {"model-uuid": model_uuid},
            {"$set": {"img.description_models": new_image_base64}}
        )

        if result.matched_count == 0:
            return jsonify({"error": "Modelo no encontrado."}), 404

        return jsonify({"message": "Imagen descriptiva cambiada con éxito."}), 200

    except Exception as e:
        logger.error(f"Error al cambiar la imagen descriptiva para {model_uuid}: {e}")
        return jsonify({"error": "Error interno del servidor."}), 500

@app.route("/admincatalogo")
def admincatalogo():
    try:
        # Crear índice en sort_order si no existe (opcional)
        catalogo_collection.create_index([("sort_order", 1)])

        # Consultar y ordenar catálogo
        catalogos = list(
            catalogo_collection.find(
                {}, {"_id": 0, "Tipo de Modelo": 1, "img": 1, "model-uuid": 1, "sort_order": 1}
            ).sort("sort_order", 1).allow_disk_use(True)
        )
        logger.info("Catálogo cargado correctamente.")
        return render_template("admin/admincatalogo.html", catalogos=catalogos)
    except Exception as e:
        flash("Error al cargar el catálogo.")
        logger.error(f"Error al cargar el catálogo: {e}")
        return redirect(url_for("login"))


# Generar cadena aleatoria de 20 caracteres
def generate_random_string(length=20):
    characters = string.ascii_letters + string.digits
    return ''.join(random.choices(characters, k=length))

@app.route("/add_catalog_item", methods=["POST"])
def add_catalog_item():
    tipo_modelo = request.form["tipo_modelo"]
    description_models_file = request.files["description_models"]
    modelos_file = request.files["modelos"]

    # Convertir imágenes a Base64
    description_models_base64 = base64.b64encode(description_models_file.read()).decode("utf-8")
    modelos_base64 = base64.b64encode(modelos_file.read()).decode("utf-8")

    model_uuid = generate_random_string()

    # Recopilar atributos únicos de Forms y corte_lazer
    existing_items = list(catalogo_collection.find())
    merged_forms = {}
    merged_corte_lazer = {}

    for item in existing_items:
        # Combinar atributos de Forms
        for key, value in item.get("Forms", {}).items():
            if key not in merged_forms:
                merged_forms[key] = value
            else:
                if isinstance(merged_forms[key], list):
                    if isinstance(value, list):
                        merged_forms[key] = list(set(merged_forms[key] + value))
                    else:
                        merged_forms[key] = list(set(merged_forms[key] + [value]))
                elif isinstance(merged_forms[key], str):
                    if isinstance(value, list):
                        merged_forms[key] = list(set([merged_forms[key]] + value))
                    else:
                        merged_forms[key] = list(set([merged_forms[key], value]))

        # Combinar atributos de corte_lazer
        for key, value in item.get("corte_lazer", {}).items():
            if key not in merged_corte_lazer:
                merged_corte_lazer[key] = value
            else:
                if isinstance(merged_corte_lazer[key], list):
                    if isinstance(value, list):
                        merged_corte_lazer[key] = list(set(merged_corte_lazer[key] + value))
                    else:
                        merged_corte_lazer[key] = list(set(merged_corte_lazer[key] + [value]))
                elif isinstance(merged_corte_lazer[key], str):
                    if isinstance(value, list):
                        merged_corte_lazer[key] = list(set([merged_corte_lazer[key]] + value))
                    else:
                        merged_corte_lazer[key] = list(set([merged_corte_lazer[key], value]))

    # Insertar nuevo elemento en la base de datos
    new_item = {
        "model-uuid": model_uuid,
        "Tipo de Modelo": tipo_modelo,
        "Forms": merged_forms,
        "corte_lazer": merged_corte_lazer,
        "img": {
            "description_models": description_models_base64,
            "modelos": modelos_base64,
        },
        "sort_order": catalogo_collection.count_documents({})  # Asigna el orden al final por defecto
    }
    catalogo_collection.insert_one(new_item)
    return redirect("/admincatalogo")

@app.route("/delete_catalog_item/<model_uuid>", methods=["DELETE"])
def delete_catalog_item(model_uuid):
    result = catalogo_collection.delete_one({"model-uuid": model_uuid})
    if result.deleted_count > 0:
        return "Deleted successfully", 200
    else:
        return "Item not found", 404


@app.route("/adminforms/<model_uuid>")
def adminforms(model_uuid):
    try:
        logger.info(f"Buscando modelo con UUID: {model_uuid}")
        model = catalogo_collection.find_one({"model-uuid": model_uuid}, {"_id": 0})
        if model:
            logger.info(f"Modelo encontrado: {model}")
            return render_template("admin/adminforms.html", model=model)
        else:
            flash("Modelo no encontrado.")
            logger.warning(f"Modelo no encontrado para UUID: {model_uuid}")
            return redirect(url_for("admincatalogo"))
    except Exception as e:
        flash("Error al cargar el formulario.")
        logger.error(f"Error al cargar el formulario: {e}")
        return redirect(url_for("admincatalogo"))
    
@app.route("/get_forms/<model_uuid>", methods=["GET"])
def get_forms(model_uuid):
    try:
        model = catalogo_collection.find_one({"model-uuid": model_uuid})
        if not model:
            return jsonify({"error": "Modelo no encontrado"}), 404

        return jsonify(model.get("Forms", {})), 200
    except Exception as e:
        logger.error(f"Error al obtener el JSON para {model_uuid}: {e}")
        return jsonify({"error": "Error interno del servidor"}), 500


@app.route("/edit_forms/<model_uuid>", methods=["POST"])
def edit_forms(model_uuid):
    try:
        updated_forms = request.json
        if not updated_forms:
            return jsonify({"error": "No se recibieron datos"}), 400

        result = catalogo_collection.update_one(
            {"model-uuid": model_uuid},
            {"$set": {"Forms": updated_forms}}
        )

        if result.matched_count == 0:
            return jsonify({"error": "Modelo no encontrado"}), 404

        return jsonify({"message": "Formulario actualizado con éxito"}), 200
    except Exception as e:
        logger.error(f"Error al guardar el JSON para {model_uuid}: {e}")
        return jsonify({"error": "Error interno del servidor"}), 500

@app.route("/get_corte_lazer/<model_uuid>", methods=["GET"])
def get_corte_lazer(model_uuid):
    """
    Endpoint para obtener el contenido de `corte_lazer` de un modelo específico.
    """
    try:
        model = catalogo_collection.find_one({"model-uuid": model_uuid}, {"_id": 0, "corte_lazer": 1})
        if not model:
            return jsonify({"error": "Modelo no encontrado"}), 404

        corte_lazer = model.get("corte_lazer", {})
        return jsonify(corte_lazer), 200
    except Exception as e:
        logger.error(f"Error al obtener corte_lazer para {model_uuid}: {e}")
        return jsonify({"error": "Error interno del servidor"}), 500
    
@app.route("/edit_corte_lazer/<model_uuid>", methods=["POST"])
def edit_corte_lazer(model_uuid):
    """
    Endpoint para actualizar el contenido de `corte_lazer` de un modelo específico.
    """
    try:
        updated_corte_lazer = request.json  # Obtener el JSON enviado en el cuerpo de la solicitud
        if not updated_corte_lazer:
            return jsonify({"error": "No se recibieron datos"}), 400

        # Actualizar el campo `corte_lazer` en la base de datos
        result = catalogo_collection.update_one(
            {"model-uuid": model_uuid},
            {"$set": {"corte_lazer": updated_corte_lazer}}
        )

        if result.matched_count == 0:
            return jsonify({"error": "Modelo no encontrado"}), 404

        return jsonify({"message": "Corte Láser actualizado con éxito"}), 200
    except Exception as e:
        logger.error(f"Error al guardar corte_lazer para {model_uuid}: {e}")
        return jsonify({"error": "Error interno del servidor"}), 500

def get_pedidos_por_dia():
    """
    Genera un gráfico de pedidos por día en el último mes.
    """
    try:
        pedidos_collection = db["pedidos"]

        # Calcular rango de fechas
        hoy = datetime.now()
        hace_un_mes = hoy - timedelta(days=30)

        # Filtrar pedidos en el último mes
        pedidos = list(pedidos_collection.find({
            "time-stamp": {"$gte": hace_un_mes.strftime("%Y-%m-%d %H:%M:%S")}
        }, {"time-stamp": 1}))

        # Contar pedidos por día
        pedidos_por_dia = {}
        for pedido in pedidos:
            time_stamp = pedido.get("time-stamp", "")
            if time_stamp:
                fecha = time_stamp.split(" ")[0]
                pedidos_por_dia[fecha] = pedidos_por_dia.get(fecha, 0) + 1

        # Ordenar fechas y cantidades
        fechas = sorted(pedidos_por_dia.keys())
        cantidades = [pedidos_por_dia[fecha] for fecha in fechas]

        # Crear gráfico con Plotly
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=fechas,
            y=cantidades,
            marker=dict(color='#1f77b4'),
            text=cantidades,
            textposition='auto'
        ))
        fig.update_layout(
            title="Pedidos por Día (Último Mes)",
            xaxis_title="Fecha",
            yaxis_title="Cantidad de Pedidos",
            xaxis=dict(tickangle=-45),
            template="plotly_white"
        )

        return json.dumps(fig, cls=PlotlyJSONEncoder)
    except Exception as e:
        logger.error(f"Error al generar gráfico de pedidos por día: {e}")
        return None

def get_top_clientes_frecuentes():
    """
    Genera un gráfico de los top 3 clientes frecuentes.
    """
    try:
        pedidos_collection = db["pedidos"]

        # Contar la cantidad de pedidos por cliente
        clientes = {}
        pedidos = pedidos_collection.find({}, {"cliente-nombre": 1})
        for pedido in pedidos:
            cliente_nombre = pedido.get("cliente-nombre", "Desconocido")
            clientes[cliente_nombre] = clientes.get(cliente_nombre, 0) + 1

        # Ordenar clientes por número de pedidos y tomar los top 3
        top_clientes = sorted(clientes.items(), key=lambda x: x[1], reverse=True)[:3]
        nombres = [cliente[0] for cliente in top_clientes]
        cantidades = [cliente[1] for cliente in top_clientes]

        # Crear gráfico con Plotly
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=nombres,
            y=cantidades,
            marker=dict(color='#ff7f0e'),
            text=cantidades,
            textposition='auto'
        ))
        fig.update_layout(
            title="Top 3 Clientes Frecuentes",
            xaxis_title="Cliente",
            yaxis_title="Cantidad de Pedidos",
            template="plotly_white"
        )

        return json.dumps(fig, cls=PlotlyJSONEncoder)
    except Exception as e:
        logger.error(f"Error al generar gráfico de top clientes frecuentes: {e}")
        return None

def get_top_modelos_mas_pedidos():
    """
    Genera un gráfico de los top 3 modelos más pedidos.
    """
    try:
        pedidos_collection = db["pedidos"]

        # Contar la cantidad total de pedidos por modelo
        modelos = {}
        pedidos = pedidos_collection.find({}, {"pedidos": 1})
        for pedido in pedidos:
            for item in pedido.get("pedidos", []):
                modelo = item.get("modelo", "Desconocido")
                cantidad = item.get("cantidad", 0)
                modelos[modelo] = modelos.get(modelo, 0) + cantidad

        # Ordenar modelos por cantidad de pedidos y tomar los top 3
        top_modelos = sorted(modelos.items(), key=lambda x: x[1], reverse=True)[:3]
        nombres = [modelo[0] for modelo in top_modelos]
        cantidades = [modelo[1] for modelo in top_modelos]

        # Crear gráfico con Plotly
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=nombres,
            y=cantidades,
            marker=dict(color='#2ca02c'),
            text=cantidades,
            textposition='auto'
        ))
        fig.update_layout(
            title="Top 3 Modelos Más Pedidos",
            xaxis_title="Modelo",
            yaxis_title="Cantidad de Pedidos",
            template="plotly_white"
        )

        return json.dumps(fig, cls=PlotlyJSONEncoder)
    except Exception as e:
        logger.error(f"Error al generar gráfico de top modelos más pedidos: {e}")
        return None
    
@app.route("/admindashboard", methods=["GET"])
def admindashboard():
    """
    Renderiza el dashboard de administrador con gráficos.
    """
    try:
        graph_pedidos_por_dia = get_pedidos_por_dia()
        graph_top_clientes = get_top_clientes_frecuentes()
        graph_top_modelos = get_top_modelos_mas_pedidos()

        if not graph_pedidos_por_dia or not graph_top_clientes or not graph_top_modelos:
            flash("Ocurrió un error al generar los gráficos.")
            return redirect(url_for("adminplatform"))

        return render_template("admin/admindashboard.html", 
                               graph_pedidos_por_dia=graph_pedidos_por_dia,
                               graph_top_clientes=graph_top_clientes,
                               graph_top_modelos=graph_top_modelos)
    except Exception as e:
        logger.error(f"Error al cargar el dashboard de administrador: {e}")
        flash("Ocurrió un error al cargar el dashboard.")
        return redirect(url_for("adminplatform"))



if __name__ == "__main__":
    debug_mode = os.getenv("DEBUG_MODE", "false").lower() == "true"  # Detecta la variable de entorno
    port = int(os.environ.get("PORT", 5000))
    
    if debug_mode:
        logger.info("El servidor está ejecutándose en modo DEBUG.")
    else:
        logger.info("El servidor está ejecutándose en modo PRODUCCIÓN.")
    
    app.run(host="0.0.0.0", port=port, debug=debug_mode)
